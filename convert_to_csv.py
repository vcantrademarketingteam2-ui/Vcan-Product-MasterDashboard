"""
VCAN Product Master — Excel to CSV + Google Sheets Updater
Reads MASTER sheet from Product Master 2026.xlsx,
exports a local CSV, and uploads to Google Sheets automatically.

Requirements:
    pip install openpyxl gspread

Usage:
    python convert_to_csv.py
"""

import openpyxl
import csv
import os
import sys
from datetime import datetime

import gspread
from google.oauth2.service_account import Credentials

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH    = r"Y:\MARKETING\Product Master\Product Master 2026.xlsx"
OUTPUT_DIR    = r"Y:\MARKETING\Product Master\ทดสอบ"
CREDENTIALS   = os.path.join(BASE_DIR, "Credentials.json", "vcan-product-master-bd8b479ba665.json")

# ── Google Sheets ──────────────────────────────────────────────────────────
SPREADSHEET_ID = "10iDgfJ8o37XS9jEjM-o8NCap_3gVNwZfv91B18riFTA"
EXCEL_SHEET_NAME = "MASTER"   # tab name in Excel file
GSHEET_TAB_NAME  = "Master"   # tab name in Google Sheets

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Barcode columns (0-indexed) — keep as strings to preserve leading zeros
BARCODE_COLS = {2, 3}


def cell_to_str(val, col_index: int) -> str:
    if val is None:
        return ""
    if col_index in BARCODE_COLS:
        if isinstance(val, (int, float)):
            return str(int(val))
        return str(val).strip()
    if col_index == 6:  # RSP column
        try:
            return f" {float(val):.2f} "
        except (TypeError, ValueError):
            return str(val).strip()
    if isinstance(val, float):
        return str(int(val)) if val == int(val) else f"{val:.2f}"
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def read_excel():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel file not found:\n  {EXCEL_PATH}")
        sys.exit(1)

    print(f"[1/3] Reading Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    if EXCEL_SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] Sheet '{EXCEL_SHEET_NAME}' not found. Available: {wb.sheetnames}")
        wb.close()
        sys.exit(1)

    ws = wb[EXCEL_SHEET_NAME]
    rows = [[cell_to_str(v, i) for i, v in enumerate(row)] for row in ws.iter_rows(values_only=True)]
    wb.close()
    print(f"      {len(rows)} rows loaded")
    return rows


def save_csv(rows):
    print(f"[2/3] Saving CSV...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_str    = datetime.now().strftime("%d-%m-%y")
    output_path = os.path.join(OUTPUT_DIR, f"Product Master 2026 {date_str}.csv")

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)

    print(f"      Saved: {output_path}")
    return output_path


def upload_to_sheets(rows):
    print(f"[3/3] Uploading to Google Sheets...")

    if not os.path.exists(CREDENTIALS):
        print(f"[ERROR] credentials.json not found:\n  {CREDENTIALS}")
        sys.exit(1)

    creds  = Credentials.from_service_account_file(CREDENTIALS, scopes=SCOPES)
    client = gspread.authorize(creds)

    sh = client.open_by_key(SPREADSHEET_ID)
    ws = sh.worksheet(GSHEET_TAB_NAME)

    ws.clear()
    ws.update(rows, value_input_option="USER_ENTERED")

    print(f"      Done! Google Sheets updated ✓")
    print(f"      https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}")


if __name__ == "__main__":
    rows = read_excel()
    save_csv(rows)
    upload_to_sheets(rows)
    print()
    print("=== All done! Dashboard will update within 1-2 minutes ===")
