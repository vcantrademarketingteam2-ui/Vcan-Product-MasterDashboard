#!/usr/bin/env python3
"""
convert_promo.py — Parse promotion plan xlsx files into src/promo_data.js
Usage: python convert_promo.py
Reads from Y:\\MARKETING\\Promotion Plan ทุกห้าง\\2026\\
Writes to src/promo_data.js
"""
import json
from datetime import datetime
from pathlib import Path
import openpyxl

BASE_DIR = Path(r'Y:\MARKETING\Promotion Plan ทุกห้าง\2026')
OUT = Path('src/promo_data.js')

# Fill colours (openpyxl ARGB AARRGGBB) grouped into families. The *meaning* of a
# family differs per retailer (Tops yellow=media but Villa yellow=field — the stores
# literally swap them), so we map hex -> family here and family -> activity per store.
COLOR_FAMILY = {
    'FFFFFF00': 'yellow', 'FFFFC000': 'yellow', 'FFFFFF99': 'yellow',
    'FFC65911': 'orange', 'FFFF9933': 'orange', 'FFED7D31': 'orange',
    'FF99FF99': 'green', 'FF92D050': 'green', 'FF00B050': 'green',
    'FFCE60D1': 'purple', 'FFE3A5E5': 'purple',
}
# Clearance is ONLY the bright cyan-blue FF00B0F0 (it paints a whole row for a clearance
# item). The lighter blues (FFA3DBFF/FF76E3FF/FF33CCCC) are just price highlights, NOT
# clearance — grouping them caused false CLEAR badges (e.g. L'Arbre Vert).
CLEARANCE_COLORS = {'FF00B0F0'}
DISCON_COLORS = {'FFFF0000'}
# Ignore: pink/grey header highlighting + the light-blue price highlights
IGNORE_COLORS = {'FFFF99FF', 'FFFF9999', 'FFD9D9D9', 'FFBDD7EE', 'FFE2EFDA',
                 'FFA3DBFF', 'FF76E3FF', 'FF33CCCC', 'FFBDD7EE'}

# Global default: orange = ลงพื้นที่ (field), yellow = ลงสื่อ (media).
ACTIVITY_BY_FAMILY = {'yellow': 'media', 'orange': 'field', 'purple': 'media', 'green': 'field'}

# Per-retailer overrides for stores that use colours differently.
# Only supply keys that differ from the global default above.
RETAILER_ACTIVITY_OVERRIDE = {
    'TWD': {'orange': 'media'},  # TWD: FFC65911 = ลงสื่อ, not field. No field activity in TWD.
}

LOOKS_KEYWORDS = ['looks', 'magazine']
SKIP_HEADER_KEYWORDS = {'total', 'remark', 'note', 'หมายเหตุ', 'grand total', 'sub total'}

# Brands sold WITHOUT 7% VAT (pet food). Their RSP is already net, so dividing by
# 1.07 understates ex-VAT price and overstates the compensate. Matched by barcode via
# the master data.js brand (promo rows have blank brand), prefix so all VITAKRAFT
# sub-brands (CAT/DOG/Bird/Rabbit/Rodent) qualify.
NON_VAT_BRAND_PREFIXES = ('VITAKRAFT',)


def load_non_vat_barcodes():
    """Read src/data.js (generated first by convert_to_data.py) and return the set of
    barcodes whose master brand is a non-VAT brand. The array is valid JSON."""
    try:
        txt = Path('src/data.js').read_text(encoding='utf-8')
        start = txt.index('[', txt.index('PRODUCT_DATA'))
        end = txt.rindex(']') + 1
        items = json.loads(txt[start:end])
    except Exception as e:
        print(f'  (!)  Could not load data.js for non-VAT brands: {e}')
        return set()
    s = {str(it.get('barcode', '')) for it in items
         if str(it.get('brand', '')).upper().startswith(NON_VAT_BRAND_PREFIXES)}
    return s

# Per-retailer config: file name, sheets, and which row the header is on (1-indexed)
# Adjust header_row and column indices if the script misreads a file.
CONFIG = {
    'Tops': {
        'file': 'Tops - Activity Promotion 2026.xlsx',
        'sheets': ['VCAN', 'อาหารสัตว์'],
        'header_row': 9,
        'cols': {'barcode': 1, 'product': 2, 'pack': 3, 'cost': 4, 'rsp': 5, 'gp': 6, 'period_start': 7},
    },
    'Villa': {
        'file': 'Villa - Activity Promotion 2026.xlsx',
        'sheets': ['Plan Update', 'Moola'],
        'header_row': 8,
        # Both Villa sheets put period NAMES (Monthly1..N) on the row directly above the
        # Barcode row (Plan Update: names r7/barcode r8; Moola: names r6/barcode r7). Their
        # date rows are stale 2022/2023 template values, so we show the Monthly label only.
        'period_names_above': True,
        # A=Barcode, B=Item code, C=Product, D=Pack, E=Cost, F=RSP, G=GP%, H+=periods
        'cols': {'barcode': 1, 'product': 3, 'pack': 4, 'cost': 5, 'rsp': 6, 'gp': 7, 'period_start': 8},
    },
    'Big C': {
        'file': 'Big C - Activity Promotion 2026.xlsx',
        'sheets': None,  # first sheet only
        'header_row': 7,
        # Columns are spread out: A=Barcode, B=Product, C=Pack, D=Cost, G=RSP, H=ปรับราคา,
        # J=GP%, then periods Bro1.. start at L. (RSP/GP are NOT at E/F like other files.)
        'cols': {'barcode': 1, 'product': 2, 'pack': 3, 'cost': 4, 'rsp': 7, 'gp': 10, 'period_start': 12},
    },
    'Boots': {
        'file': 'Boots Plan promotion  2026.xlsx',
        'sheets': ['Sheet1'],
        'header_row': 3,
        # Col A empty, B=Barcode, C=Item code, D=Product, E=Pack, F=Cost, G=RSP, H=GP%, I+=periods
        'cols': {'barcode': 2, 'product': 4, 'pack': 5, 'cost': 6, 'rsp': 7, 'gp': 8, 'period_start': 9},
    },
    'Foodland': {
        'file': 'Foodland - Activity Promotion 2026.xlsx',
        'sheets': None,
        'header_row': 7,
        # A=Barcode, B=Code Foodland, C=Product, D=Pack, E=Cost, F=RSP, G=GP%, H+=periods
        'cols': {'barcode': 1, 'product': 3, 'pack': 4, 'cost': 5, 'rsp': 6, 'gp': 7, 'period_start': 8},
    },
    'Homepro': {
        'file': 'Homepro - Activity Promotion 2026.xlsx',
        'sheets': None,
        'header_row': 6,
        # A=Barcode, B=Article, C=Product, D=Pack, E=Cost, F=RSP, G=GP%, H+=periods
        'cols': {'barcode': 1, 'product': 3, 'pack': 4, 'cost': 5, 'rsp': 6, 'gp': 7, 'period_start': 8},
    },
    'Lotus': {
        'file': 'Lotus - Activity Promotion 2026.xlsx',
        'sheets': None,
        'header_row': 7,
        # A=Barcode, B=TPNA item code, C=Product, D=Pack size, E=Cost, F=RSP, G=GP%, H+=periods
        # (was off-by-one: read the TPNA code as the product name, leaking it into brand pills)
        'cols': {'barcode': 1, 'product': 3, 'pack': 4, 'cost': 5, 'rsp': 6, 'gp': 7, 'period_start': 8},
    },
    'TWD': {
        'file': 'TWD - Activity Promotion 2026.xlsx',
        'sheets': ['TWD'],
        'header_row': 5,
        # Columns alternate month + brochure: Jan, Bro.1, Feb, Bro.2 ... A value in a 'Bro.N'
        # column = a media (brochure) placement, so treat those as media regardless of colour.
        'bro_is_media': True,
        # A=Barcode, B=Item Code, C=Product, D=Pack, E=Cost, F=RSP, G=GP%, H+=periods
        'cols': {'barcode': 1, 'product': 3, 'pack': 4, 'cost': 5, 'rsp': 6, 'gp': 7, 'period_start': 8},
    },
    # Central: add path when available
    # 'Central Department': {
    #     'file': 'Central - Activity Promotion 2026.xlsx',
    #     'sheets': None,
    #     'header_row': 8,
    #     'cols': {'barcode': 1, 'product': 2, 'pack': 3, 'cost': 4, 'rsp': 5, 'gp': 6, 'period_start': 7},
    # },
}


def is_barcode(val):
    s = str(val or '').strip().split('.')[0]  # handle float barcodes
    return s.isdigit() and len(s) >= 8


def fmt_date_cell(val):
    """Render a date-range cell as a compact single-line string. Excel often stores these
    as datetime objects which str() turns into '2022-11-08 00:00:00' (long, wraps) — collapse
    those to 'DD/MM/YY'. Plain strings are just whitespace-collapsed."""
    from datetime import datetime, date
    if isinstance(val, (datetime, date)):
        return val.strftime('%d/%m/%y')
    return ' '.join(str(val).split())  # collapse internal whitespace/newlines


def get_fill_color(cell):
    """Return ARGB hex string if cell has a solid fill, else None."""
    try:
        fill = cell.fill
        if fill and fill.fill_type == 'solid':
            c = fill.fgColor
            if c.type == 'rgb':
                rgb = c.rgb.upper()
                if rgb not in ('00000000', 'FFFFFFFF', 'FF000000'):
                    return rgb
    except Exception:
        pass
    return None


def detect_marks(cell, retailer):
    """Inspect one promo cell and return (activities, is_blue, has_comment).
      - activities: list of 'media'/'field'/'looks' (LOOKS from a comment, others via the
        global colour map).
      - is_blue: True if FF00B0F0-filled (whole-row clearance flag, accumulated per product).
      - has_comment: True if the cell carries any comment (so we don't drop a LOOKS note that
        sits on an empty price cell).
    """
    acts = []
    is_blue = False
    has_comment = False
    try:
        if cell.comment:
            has_comment = True
            if 'look' in cell.comment.text.lower():
                acts.append('looks')
    except Exception:
        pass
    color = get_fill_color(cell)
    if color and color not in IGNORE_COLORS:
        if color in DISCON_COLORS:
            pass  # discontinued is surfaced at product level via master status, not here
        elif color in CLEARANCE_COLORS:
            is_blue = True
        else:
            fam = COLOR_FAMILY.get(color)
            if fam:
                override = RETAILER_ACTIVITY_OVERRIDE.get(retailer, {})
                act = override.get(fam, ACTIVITY_BY_FAMILY.get(fam))
                if act and act not in acts:
                    acts.append(act)
    return acts, is_blue, has_comment


def num(val):
    """Safely convert to float or None."""
    try:
        return float(val) if val is not None else None
    except (ValueError, TypeError):
        return None


def detect_header_row(rows, cfg_fallback):
    """Find the row index (0-based) containing 'Barcode' in the first few columns.
    Sheets in the same file can have headers on different rows (e.g. Tops VCAN row 9
    vs อาหารสัตว์ row 3), so we detect per-sheet instead of trusting one config value."""
    for ri, row in enumerate(rows[:20]):
        for ci in range(0, 4):
            if ci < len(row):
                v = row[ci].value
                if v is not None and str(v).strip().lower() == 'barcode':
                    return ri
    return cfg_fallback - 1  # fall back to config (1-indexed -> 0-indexed)


def looks_like_period(name, date_range):
    """Real periods are dates or month/cycle labels ('2603/04', 'Jan', 'SP6'). Junk
    columns from mis-structured sheets hold promo *values* ('75', 'buy 2 get 1',
    'Buy2 145') — reject pure numbers and promo-keyword strings."""
    n = str(name).strip().lower()
    if not n:
        return False
    # pure number (with optional decimal/comma) = a promo price, not a period
    if n.replace('.', '').replace(',', '').isdigit():
        return False
    # promo-deal wording leaked into a header
    if any(k in n for k in ('buy', 'get', 'free', 'แถม', 'ซื้อ')):
        return False
    return True


def parse_sheet(ws, cfg, retailer, non_vat=frozenset()):
    """Parse one worksheet, return (products list, periods list)."""
    rows = list(ws.iter_rows(values_only=False))
    h = detect_header_row(rows, cfg['header_row'])
    if h >= len(rows):
        print(f'    (!)  Header row out of range (sheet has {len(rows)} rows)')
        return [], []

    c = cfg['cols']
    ps = c['period_start'] - 1  # 0-indexed

    # Period-name row location varies. Villa's two sheets both put the period NAMES
    # (Monthly1..N) on the row directly above the Barcode row — and detect_header_row
    # finds that Barcode row per sheet — so 'period_names_above' = h - 1 works for both,
    # unlike a fixed row number which only matched one sheet.
    if cfg.get('period_names_above'):
        name_row_i = h - 1
        # Date range sits 2 rows above the period-name row on Villa's "Plan Update" (VCAN)
        # sheet → real 2026 ranges. The "Moola" sheet is offset differently, so this lands on
        # an empty row there, dodging its stale 2022/23 template dates. Plan Update is parsed
        # first and the period union keeps first-seen, so Villa gets the good dates.
        # ponytail: row-offset coincidence; if Villa's sheet layout shifts, re-check this.
        date_row_i = name_row_i - 2
    else:
        name_row_i = (cfg['period_name_row'] - 1) if cfg.get('period_name_row') else h
        date_row_i = (cfg['period_date_row'] - 1) if cfg.get('period_date_row') else h + 1
    name_row = rows[name_row_i] if 0 <= name_row_i < len(rows) else []
    date_row = rows[date_row_i] if (date_row_i is not None and date_row_i < len(rows)) else []

    # Extract period column definitions
    periods = []
    seen_names = set()
    for ci in range(ps, len(name_row)):
        cell = name_row[ci]
        val = cell.value
        if not val:
            continue
        name = str(val).strip()
        if not name or name.lower() in SKIP_HEADER_KEYWORDS:
            continue
        date_range = ''
        if ci < len(date_row) and date_row[ci].value:
            date_range = fmt_date_cell(date_row[ci].value)
        if not looks_like_period(name, date_range):
            continue  # skip junk columns (promo values, calc columns)
        if name in seen_names:
            name = f'{name}_{ci}'  # deduplicate
        seen_names.add(name)
        periods.append({'name': name, 'dateRange': date_range, 'colIdx': ci})

    # Pre-scan: find which (period_col_idx, data_row_offset) pairs carry a LOOKS comment.
    # Planners often note "LOOKs Magazine" only on the first row of a product group; the
    # other variants in the same group just have the media fill colour. We propagate LOOKS
    # to any row within LOOKS_PROXIMITY of a LOOKS-commented cell in the same period column.
    LOOKS_PROXIMITY = 25  # rows
    looks_by_col = {}  # col_idx -> [row_offsets with LOOKS comment]
    for ri_off, row in enumerate(rows[h + 2:]):
        for p in periods:
            ci = p['colIdx']
            if ci < len(row):
                try:
                    c2 = row[ci]
                    if c2.comment and 'look' in c2.comment.text.lower():
                        looks_by_col.setdefault(ci, []).append(ri_off)
                except Exception:
                    pass

    def near_looks(ri_off, col_idx):
        for r in looks_by_col.get(col_idx, []):
            if abs(r - ri_off) <= LOOKS_PROXIMITY:
                return True
        return False

    # Parse product rows (skip header + date rows). Track the real 1-based row number so we
    # can skip rows the planner hid in Excel (hidden = intentionally excluded from the plan).
    products = []
    for ri_off, row in enumerate(rows[h + 2:]):
        row_num = h + 2 + ri_off + 1  # 1-based Excel row
        if ws.row_dimensions[row_num].hidden:
            continue
        if len(row) <= c['barcode'] - 1:
            continue
        bc_cell = row[c['barcode'] - 1]
        bc_val = bc_cell.value if bc_cell else None
        if not is_barcode(bc_val):
            continue

        # Normalize barcode
        barcode = str(int(float(str(bc_val).strip()))) if bc_val else ''
        if len(barcode) == 11:
            barcode = '0' + barcode
        if len(barcode) < 8:
            continue

        prod_cell = row[c['product'] - 1] if c['product'] - 1 < len(row) else None
        product = str(prod_cell.value or '').strip() if prod_cell else ''
        # Remove non-breaking spaces (\xa0), soft-hyphens (\xad), and other invisible
        # Unicode control/format characters that appear in some Excel cells (e.g. Combat·name)
        product = product.replace('\xa0', ' ').replace('\xad', '').replace('​', '').strip()
        product = ' '.join(product.split())  # collapse internal whitespace
        if not product or product.lower() in ('total', 'grand total'):
            continue

        pack_cell = row[c['pack'] - 1] if c['pack'] - 1 < len(row) else None
        pack = str(pack_cell.value or '').strip() if pack_cell else ''

        cost_cell = row[c['cost'] - 1] if c['cost'] - 1 < len(row) else None
        cost = num(cost_cell.value if cost_cell else None)

        # Non-VAT (pet food) RSP is already net — don't strip a 7% that isn't there.
        vat_div = 1.0 if barcode in non_vat else 1.07

        rsp_cell = row[c['rsp'] - 1] if c['rsp'] - 1 < len(row) else None
        rsp_inc = num(rsp_cell.value if rsp_cell else None)
        rsp_ex = round(rsp_inc / vat_div, 4) if rsp_inc else None

        gp_cell = row[c['gp'] - 1] if c['gp'] - 1 < len(row) else None
        gp_raw = num(gp_cell.value if gp_cell else None)
        # GP stored as decimal (0.35) or percent (35) — normalize to decimal
        gp = gp_raw if gp_raw is not None and gp_raw <= 1.0 else (gp_raw / 100 if gp_raw else None)

        # Parse period cells
        period_data = {}
        any_blue = False  # blue paints the whole row -> product is a clearance item
        bro_is_media = cfg.get('bro_is_media')
        for p in periods:
            ci = p['colIdx']
            if ci >= len(row):
                continue
            cell = row[ci]
            val = cell.value
            blank = val is None or (isinstance(val, str) and not val.strip())

            activities, is_blue, has_comment = detect_marks(cell, retailer)
            if is_blue:
                any_blue = True
            # TWD encodes media as the brochure ('Bro.N') columns themselves — a value in one
            # means the SKU is placed in that brochure, regardless of fill colour.
            is_bro = bro_is_media and 'bro' in str(p['name']).lower()
            if is_bro and not blank and 'media' not in activities:
                activities.append('media')

            # LOOKS propagation: planners note "LOOKs Magazine" on ONE cell per group then
            # use plain media fill on the other rows.  If this cell has 'media' and a LOOKS
            # comment exists in the same period column within LOOKS_PROXIMITY rows, upgrade
            # the media activity to 'looks'.  Brochure-based media (Bro.N columns) is
            # excluded — it's a different kind of media placement.
            if 'media' in activities and 'looks' not in activities and not is_bro:
                if near_looks(ri_off, p['colIdx']):
                    activities = ['looks' if a == 'media' else a for a in activities]

            # Keep the cell only if it carries something to show: a price, an activity
            # (incl. a LOOKS note that sits on an empty cell), or it's a brochure placement.
            if blank and not activities and not (is_bro and has_comment):
                continue

            sale_price = None
            sale_label = None
            if not blank:
                if isinstance(val, (int, float)):
                    sale_price = float(val)
                else:
                    v = num(str(val).replace(',', ''))
                    sale_price = v if v is not None else None
                    if v is None:
                        sale_label = str(val).strip()  # e.g. "Buy2Get1", "3for499"

            # Compensate = RSP_ex - (sale_price_inc / vat_div)
            compensate = None
            if sale_price is not None and rsp_ex is not None:
                compensate = round(rsp_ex - (sale_price / vat_div), 4)

            period_data[p['name']] = {
                'salePrice': sale_price,
                'saleLabel': sale_label,
                'activities': activities,
                'compensate': compensate,
            }

        if period_data:  # only include rows that have at least one promotional period
            products.append({
                'retailer': retailer,
                'barcode': barcode,
                'product': product,
                'brand': '',   # filled in by App.jsx via rawData barcode lookup
                'company': '',
                'packSize': pack,
                'rspExVat': rsp_ex,
                'rspIncVat': rsp_inc,
                'cost': cost,
                'gp': gp,
                'clearance': any_blue,  # product-level clearance flag
                'periods': period_data,
            })

    return products, periods


def main():
    all_products = []
    promo_meta = {}
    promo_retailers = []

    non_vat_barcodes = load_non_vat_barcodes()
    print(f'  Non-VAT barcodes loaded: {len(non_vat_barcodes)}')
    assert non_vat_barcodes, 'No non-VAT barcodes found — data.js missing/changed? (Vitakraft lookup failed)'

    for retailer, cfg in CONFIG.items():
        fpath = BASE_DIR / cfg['file']
        if not fpath.exists():
            print(f'  (!)  Skipping {retailer}: not found at {fpath}')
            continue

        print(f'  Parsing {retailer}...')
        try:
            wb = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
        except Exception as e:
            print(f'    FAIL Failed to open: {e}')
            continue

        sheets = cfg.get('sheets') or [wb.sheetnames[0]]
        retailer_periods = []
        retailer_products = []

        by_barcode = {}  # dedup within a retailer: same barcode -> merge period dicts
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                # Try first sheet as fallback
                print(f'    (!)  Sheet "{sheet_name}" not found — trying first sheet')
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            products, periods = parse_sheet(ws, cfg, retailer, non_vat_barcodes)
            for prod in products:
                bc = prod['barcode']
                if bc in by_barcode:
                    # merge: keep existing fields, fold in any periods it doesn't have yet
                    for pname, pdata in prod['periods'].items():
                        by_barcode[bc]['periods'].setdefault(pname, pdata)
                else:
                    by_barcode[bc] = prod
                    retailer_products.append(prod)
            # Merge periods (union, preserve order)
            existing = {p['name'] for p in retailer_periods}
            for p in periods:
                if p['name'] not in existing:
                    retailer_periods.append({'name': p['name'], 'dateRange': p['dateRange']})
                    existing.add(p['name'])

        wb.close()

        if retailer_products:
            all_products.extend(retailer_products)
            promo_meta[retailer] = {'periods': retailer_periods}
            promo_retailers.append(retailer)
            print(f'    OK {len(retailer_products)} products, {len(retailer_periods)} periods')
        else:
            print(f'    (!)  No promo products found (check header_row in CONFIG)')

    # Write promo_data.js
    now = datetime.now().isoformat()
    js_lines = [
        '// Auto-generated by convert_promo.py — do not edit by hand.',
        f'// Generated: {now}',
        '',
        f'export const GENERATED_AT = {json.dumps(now)};',
        '',
        'export const PROMO_ACTIVITY = {',
        "  field: { label: 'ลงพื้นที่', color: '#06B6D4' },",
        "  media: { label: 'ลงสื่อ', color: '#A855F7' },",
        "  looks: { label: 'LOOKS Magazine', color: '#E879F9' },",
        '};',
        '',
        f'export const PROMO_RETAILERS = {json.dumps(promo_retailers)};',
        '',
        'export const PROMO_META = ' + json.dumps(promo_meta, ensure_ascii=False, indent=2) + ';',
        '',
        'const promoData = ' + json.dumps(all_products, ensure_ascii=False, indent=2) + ';',
        '',
        'export default promoData;',
    ]

    OUT.write_text('\n'.join(js_lines), encoding='utf-8')
    print(f'\nDone -- {len(all_products)} records across {len(promo_retailers)} retailers -> {OUT}')
    if not all_products:
        print('   Tip: If all retailers were skipped, check BASE_DIR path and CONFIG header_row values.')


if __name__ == '__main__':
    main()
