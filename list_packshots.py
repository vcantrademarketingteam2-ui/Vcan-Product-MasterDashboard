import os
import openpyxl

EXCEL_PATH = r"Y:\MARKETING\Meth\ห้องทำงาน Claude\product code ,name.xlsx"
FRONT_FOLDER = r"Y:\MARKETING\Meth\ห้องทำงาน Claude\Product Packshot for dashboard\Sundae\SUNDAE Pack Shot Front"
BACK_FOLDER  = r"Y:\MARKETING\Meth\ห้องทำงาน Claude\Product Packshot for dashboard\Sundae\Back"

def list_files(folder):
    return sorted(
        f for f in os.listdir(folder)
        if os.path.isfile(os.path.join(folder, f))
        and f.lower() != "thumbs.db"
    )

front_files = list_files(FRONT_FOLDER)
back_files  = list_files(BACK_FOLDER)

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.worksheets[0]  # Sheet 1 (0-indexed)

# Write headers
ws["D1"] = "Front Packshot"
ws["E1"] = "Back Packshot"

# Write filenames starting from row 2
for i, name in enumerate(front_files, start=2):
    ws.cell(row=i, column=4, value=name)

for i, name in enumerate(back_files, start=2):
    ws.cell(row=i, column=5, value=name)

wb.save(EXCEL_PATH)

print(f"Done — wrote {len(front_files)} front files to col D, {len(back_files)} back files to col E")
print("\nFront files:")
for f in front_files:
    print(f"  {f}")
print("\nBack files:")
for f in back_files:
    print(f"  {f}")
