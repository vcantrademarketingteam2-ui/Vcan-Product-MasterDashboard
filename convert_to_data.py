"""
convert_to_data.py
อ่าน Product Master 2026.xlsx แล้ว generate:
  src/data.js          - ข้อมูลสินค้าทั้งหมด (แทน Google Sheets CSV)
  src/retailer_data.js - Cost/Unit, Cost/Case, GP% ต่อบาร์โค้ดต่อห้าง

รัน: py convert_to_data.py
"""
import openpyxl, json, re
from datetime import datetime
from pathlib import Path

XLSX_PATH = r"Y:\MARKETING\Product Master\Product Master 2026.xlsx"
SRC_DIR   = Path(__file__).parent / "src"

# Barcode -> brand, for rows where the MASTER sheet leaves the brand cell blank
# (e.g. the very first product row has no brand row above it to inherit from).
BRAND_FALLBACK = {
    "3450601046278": "L'Arbre Vert",
}

RETAILERS = [
    'Tops', 'Villa', 'The Mall', 'Lotus', 'Homepro', 'Big C',
    'TWD', 'Boots', 'Foodland', 'Central Department', "Pet'n me", 'Fuji'
]

# sheet tab name -> dashboard retailer name
SHEET_MAP = {
    'TOPS'        : 'Tops',
    'Villa'       : 'Villa',
    'The mall'    : 'The Mall',
    'Lotus'       : 'Lotus',
    'HOMEPRO'     : 'Homepro',
    'HomePro WMF' : 'Homepro',
    'BIG C'       : 'Big C',
    'TWD'         : 'TWD',
    'Foodland'    : 'Foodland',
    'Boots'       : 'Boots',
    'central WMF' : 'Central Department',
    "Pet'n me"    : "Pet'n me",
}


# helpers
def tr(v):
    return str(v).strip() if v is not None else ''

def is_bc(s):
    s = tr(s).split('.')[0]
    return s.isdigit() and len(s) >= 8

def fix_bc(s):
    s = tr(s).split('.')[0]
    if len(s) == 11:   # Google Sheets / Excel strips UPC-A leading zero
        s = '0' + s
    return s

def to_num(v):
    if v is None:
        return None
    try:
        return round(float(str(v).replace('%','').replace('฿','').replace(',','').strip()), 4)
    except Exception:
        return None

def find_col(header, *keywords):
    for i, h in enumerate(header):
        hl = tr(h).lower()
        if any(kw.lower() in hl for kw in keywords):
            return i
    return None


# MASTER parser
def parse_master(ws):
    VENDORS = ['Vcan', 'Moola']

    def norm_brand(b):
        b = re.sub(r'\s+', ' ', str(b).replace('\n', ' ').replace('\r', ' ')).strip()
        if re.search(r'dove.?men', b, re.I) or b.lower() == 'shampoo':
            return 'Dove Men'
        return b

    # find header row (row where col0 == 'company')
    hrow = None
    for i, row in enumerate(ws.iter_rows(values_only=True, max_col=25)):
        if tr(row[0]).lower() == 'company':
            hrow = i
            break
    if hrow is None:
        raise ValueError("MASTER: header row not found")

    products, company, brand = [], '', ''

    for row in ws.iter_rows(min_row=hrow + 2, values_only=True, max_col=25):
        row = [tr(c) for c in row]
        if not any(row):
            continue
        c0, c1, c2 = row[0], row[1], row[2]

        if c0 in VENDORS:
            company = c0
            # A vendor name often shares its row with the section's FIRST product
            # (e.g. "Vcan | VITAKRAFT CAT | <barcode> | ..."). Skip only pure header
            # rows; otherwise fall through so the brand + product on this row are kept.
            if not (is_bc(c1) or is_bc(c2)):
                continue

        barcode = product = pack_size = rsp_raw = status = ''
        roff = 8

        if is_bc(c1) and c0 and c0 not in VENDORS and not is_bc(c0):
            brand = norm_brand(c0)
            barcode, product, pack_size = c1, row[3], row[4]
            rsp_raw, status, roff = row[5], row[6], 7
        elif is_bc(c2):
            if c1 and not is_bc(c1):
                brand = norm_brand(c1)
            barcode, product, pack_size = c2, row[4], row[5]
            rsp_raw, status, roff = row[6], row[7], 8
        else:
            continue

        if not product or product.lower() == 'total':
            continue

        barcode = fix_bc(barcode)
        if not is_bc(barcode):
            continue

        b = brand
        if b != 'Dove Men' and re.search(r'dove.?men', product, re.I):
            b = 'Dove Men'
        if not b:
            b = BRAND_FALLBACK.get(barcode, '')

        retailers = {}
        for j, r in enumerate(RETAILERS):
            idx = roff + j
            retailers[r] = row[idx] if idx < len(row) else ''

        try:
            rsp_val = float(re.sub(r'[^\d.]', '', rsp_raw)) if rsp_raw else 0
        except Exception:
            rsp_val = 0

        products.append({
            'company'  : company,
            'brand'    : b,
            'barcode'  : barcode,
            'product'  : product,
            'packSize' : pack_size,
            'rsp'      : rsp_val,
            'status'   : status,
            'retailers': retailers,
        })

    return products


# RETAILER sheet parser
def parse_retailer(ws, sheet_name):
    """Returns {barcode: {costUnit, costCase, gp}}"""

    # Pet'n me has no 'Barcode' label — barcode is in col 0
    if sheet_name == "Pet'n me":
        result = {}
        for row in ws.iter_rows(values_only=True, max_col=10):
            bc = fix_bc(tr(row[0]))
            if not is_bc(bc):
                continue
            e = {}
            cu = to_num(row[4] if len(row) > 4 else None)
            cc = to_num(row[3] if len(row) > 3 else None)
            gp = to_num(row[6] if len(row) > 6 else None)
            if cu is not None: e['costUnit'] = cu
            if cc is not None: e['costCase'] = cc
            if gp is not None: e['gp']       = gp
            if e:
                result[bc] = e
        return result

    # All other sheets: find header row containing 'barcode'
    hrow_idx, header = None, []
    for i, row in enumerate(ws.iter_rows(values_only=True, max_col=15)):
        if 'barcode' in [tr(c).lower() for c in row]:
            hrow_idx = i
            header   = [tr(c) for c in row]
            break

    if hrow_idx is None:
        print(f"    [!] No barcode header in {sheet_name}")
        return {}

    bc_col   = find_col(header, 'barcode')
    unit_col = find_col(header, 'cost/unit', 'unit (ex')
    case_col = find_col(header, 'cost/case', 'case (ex')
    gp_col   = find_col(header, 'gp%', 'normal gp')

    result = {}
    for row in ws.iter_rows(min_row=hrow_idx + 2, values_only=True, max_col=15):
        if bc_col is None or bc_col >= len(row):
            continue
        bc = fix_bc(tr(row[bc_col]))
        if not is_bc(bc):
            continue
        e = {}
        cu  = to_num(row[unit_col] if unit_col and unit_col < len(row) else None)
        cc  = to_num(row[case_col] if case_col and case_col < len(row) else None)
        gp  = to_num(row[gp_col]   if gp_col   and gp_col   < len(row) else None)
        if cu  is not None: e['costUnit'] = cu
        if cc  is not None: e['costCase'] = cc
        if gp  is not None: e['gp']       = gp
        if e:
            result[bc] = e

    return result


def main():
    print(f"\nReading: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # 1. MASTER -> src/data.js
    print("\n[1/2] Parsing MASTER sheet...")
    products = parse_master(wb['MASTER'])
    print(f"      {len(products)} products found")

    generated_at = datetime.now().isoformat(timespec='minutes')
    data_js = (
        "// Auto-generated by convert_to_data.py — do not edit manually\n"
        f'export const GENERATED_AT = "{generated_at}"\n'
        f"const PRODUCT_DATA = {json.dumps(products, ensure_ascii=False, indent=2)}\n"
        "export default PRODUCT_DATA\n"
    )
    (SRC_DIR / 'data.js').write_text(data_js, encoding='utf-8')
    print("      -> src/data.js written")

    # 2. RETAILER SHEETS -> src/retailer_data.js
    print("\n[2/2] Parsing retailer sheets...")
    retailer_data = {}

    for sheet_name, retailer_name in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [!] '{sheet_name}' not in workbook")
            continue
        rows = parse_retailer(wb[sheet_name], sheet_name)
        print(f"  {sheet_name:<20} -> {retailer_name:<22} {len(rows)} barcodes")
        for bc, entry in rows.items():
            if bc not in retailer_data:
                retailer_data[bc] = {}
            if retailer_name not in retailer_data[bc]:
                retailer_data[bc][retailer_name] = entry

    retail_js = (
        "// Auto-generated by convert_to_data.py — do not edit manually\n"
        f"const RETAILER_DATA = {json.dumps(retailer_data, ensure_ascii=False, indent=2)}\n"
        "export default RETAILER_DATA\n"
    )
    (SRC_DIR / 'retailer_data.js').write_text(retail_js, encoding='utf-8')
    n_entries = sum(len(v) for v in retailer_data.values())
    print(f"\n  -> src/retailer_data.js written ({len(retailer_data)} barcodes, {n_entries} retailer entries)")
    print("\nAll done!")


if __name__ == '__main__':
    main()
    try:
        input("\nกด Enter เพื่อปิด...")
    except (EOFError, OSError):
        pass
